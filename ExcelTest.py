# import pandas lib as pd

import openpyxl

OPEN_TRADES_IN_EXCEL = []


def update_excel_column_value():
    workbook = openpyxl.load_workbook('your_excel_file.xlsx')
    sheet = workbook['Sheet1']
    row_number = 2
    column_number = 3
    new_value = "Updated Value"
    sheet.cell(row=row_number, column=column_number, value=new_value)
    workbook.save('your_excel_file.xlsx')


# dataframe = openpyxl.load_workbook("your_excel_file.xlsx")
# dataframe1 = dataframe.active
# row_index = 0
# for row in range(0, dataframe1.max_row):
#     col_index = 0
#     obj = {}
#     if row_index == 0:
#         row_index += 1
#         continue
#     for col in dataframe1.iter_cols(1, dataframe1.max_column):
#         # print(col[row].value)
#         if col_index == 0:
#             obj["date"] = col[row].value
#         if col_index == 1:
#             obj["symbol"] = col[row].value
#         if col_index == 2:
#             obj["operator"] = col[row].value
#         if col_index == 3:
#             obj["price"] = col[row].value
#         if col_index == 4:
#             obj["lot"] = col[row].value
#         if col_index == 5:
#             obj["status"] = col[row].value
#
#         col_index += 1
#     print(obj)
#     if obj["status"].lower() == "Pending".lower():
#         OPEN_TRADES_IN_EXCEL.append(obj)
#     row_index += 1
#
# print(OPEN_TRADES_IN_EXCEL)
#update_excel_column_value()

# Read the text file
# file_path = 'Trade_notepad.txt'
#
# with open(file_path, 'r') as file:
#     lines = file.readlines()
#
# # Modify the contents (for example, add a new line)
# new_line = "This is a new line.\n"
# lines.append(new_line)
#
# # Update the same file
# with open(file_path, 'w') as file:
#     file.writelines(lines)
#
# print("File updated successfully.")


def filter_pending_status_rows(file_path):
    pending_rows = []

    with open(file_path, 'r') as file:
        lines = file.readlines()

    header = [x.strip() for x in lines[0].split(' | ')]
    for line in lines[1:]:
        parts = [x.strip() for x in line.split(' | ')]
        row = dict(zip(header, parts))

        if row['Status'] == 'Pending':
            # Convert ID to an integer
            row['Id'] = int(row['Id'])
            # Remove 'Status' key as it's not needed in the output
            #del row['Status']
            pending_rows.append(row)

    return pending_rows


# Example usage:
file_path = 'Trade_notepad.txt'
pending_rows = filter_pending_status_rows(file_path)
print(pending_rows)



def update_status_by_id(file_path, id_to_update, new_status):
    # Read the text file
    with open(file_path, 'r') as file:
        lines = file.readlines()

    updated_lines = []

    for line in lines:
        parts = line.strip().split(' | ')
        if len(parts) >= 6:
            current_id = parts[0]
            current_status = parts[5]
            if current_id == str(id_to_update):
                parts[5] = new_status
                updated_line = ' | '.join(parts)
                updated_lines.append(updated_line + '\n')
            else:
                updated_lines.append(line)

    # Write the updated content back to the file
    with open(file_path, 'w') as file:
        file.writelines(updated_lines)

    print(f"Updated Status for ID {id_to_update} to '{new_status}'.")

# Example usage:
file_path = 'Trade_notepad.txt'
update_status_by_id(file_path, 2, 'Arun')